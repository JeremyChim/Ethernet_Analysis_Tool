from openpyxl import Workbook

wb2 = Workbook()
ws2 = wb2.create_sheet(title='UDP_MSG_146',index=0)

for i in range(2,43):
    ws2.cell(1,i-1).value = i

wb2.save('UDP_MSG_146.csv')
wb2.close()
