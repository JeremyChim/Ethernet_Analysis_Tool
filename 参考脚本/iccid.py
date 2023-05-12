import os
import re
import openpyxl
from openpyxl.styles import PatternFill
import xlrd #pip install xlrd==1.2.0
import sys
import xlsxwriter
import xlwt

#def adb_shell(cmd):
    #out = os.popen(cmd).read()
    #print(out)

path=r"D:\Program Files\JetBrains\python\7B120\ICCID2.xlsx"
workbook=openpyxl.load_workbook(path)
worksheet=workbook.worksheets[0]

result=open('ICCID.txt','r')

for i in range(47,99):
    for line in result.readlines():  # 逐行读取
        if line.strip():  # 去掉空白行
            line = re.findall(r'["](.*?)["]', line)  # 正则表达式，只保留“ ”里面的内容，以列表的形式返回
            for value in line:  # 将列表的内容输出
                # print(value)
                worksheet.cell(i, 2).value = str(value)  # 写入单元格中
                i += 1
result.close()

workbook.save('ICCID4.xlsx')
workbook.close()
