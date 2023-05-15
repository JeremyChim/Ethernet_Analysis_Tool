# -*- coding: utf-8 -*-
# @Time ： 2023/5/11 17:08
# @Auth ： JeremyChim
# @File ：udp_msg_1.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim

from openpyxl import Workbook
import openpyxl

class udp_data():

    def __init__(self, udp_data):
        self.data = udp_data
        self.id = self.data[16:24]
        self.msg = self.data[40:]

    def config(self):
        self.wb = openpyxl.load_workbook(filename='config.xlsx')
        if self.id == '00000146':
            self.ws = self.wb.worksheets[0]
            self.len = 43

    def cut(self):
        self.hex_lst = []
        for i in range(2, self.len):
            a = self.ws.cell(i, 3).value * 2
            b = self.ws.cell(i, 4).value * 2 + a
            self.hex_lst.append(self.msg[a:b])

    def go_int(self):
        self.int_lst = []
        for i in self.hex_lst:
            i = int(i, 16)
            self.int_lst.append(i)

    def calc(self):
        self.msg_lst = []
        for i in range(2, self.len):
            fa = self.ws.cell(i, 5).value
            of = self.ws.cell(i, 6).value
            j = self.int_lst[i - 2] * fa + of
            self.msg_lst.append(j)

    def title(self):
        if self.id == '00000146':
            self.wb2 = Workbook()
            self.ws2 = self.wb2.create_sheet(title='UDP_MSG_146', index=0)
            for i in range(2, self.len):
                self.ws2.cell(1, i).value = self.ws.cell(i, 1).value

    def save(self, a):
            b = a + 2
            self.ws2.cell(b, 1).value = a
            for i in range(2, self.len):
                self.ws2.cell(b, i).value = self.msg_lst[i - 2]
            self.wb2.save('UDP_MSG_146.csv')
            self.wb2.close()

    def log(self, a):
        self.config()
        self.cut()
        self.go_int()
        self.calc()
        self.title()
        # self.save(a)

if __name__ == '__main__':
    da = '0011031900000077000001465609d51b0000000000976103d0c5009cbe03cd1a00c30503cfcd0002d31e01273c271227422777274526ef09ac5e0f76640f72051d74b300000000000000000000000000001c308c0119180098b103d0b6009ccb03cd1900c32903cfcc1e83e60f42930f83e40f2ec6018002f4f17ff8eae58000247301626902bf16050bc3'
    da2 = udp_data(da)
    da2.log(1)