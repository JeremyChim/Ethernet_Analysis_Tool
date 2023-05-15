# -*- coding: utf-8 -*-
# @Time ： 2023/5/11 17:08
# @Auth ： JeremyChim
# @File ：udp_msg_1.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim

from openpyxl import Workbook
import openpyxl
import csv

class udp_data():

    def __init__(self, udp_data, row):
        self.data = udp_data
        self.id = self.data[16:24]
        self.msg = self.data[40:]
        self.row = row

    def cr_title(self):
        self.config()
        self.title()

    def log(self):
        if self.id == '00000146':
            self.config()
            self.cut()
            self.go_int()
            self.calc()
            self.op()
            self.wri()
            self.save()
        elif self.id == '0000017f':
            self.config()
            self.cut()
            self.go_int()
            self.calc()
            self.op()
            self.wri()
            self.save()

    def config(self):
        self.wb = openpyxl.load_workbook(filename='../function/config/config.xlsx')
        if self.id == '00000146':
            self.ws = self.wb.worksheets[0]
            self.len = 43
        elif self.id == '0000017f':
            self.ws = self.wb.worksheets[1]
            self.len = 33
        # elif self.id == '00000271':
        #     self.ws = self.wb.worksheets[2]
        #     self.len = 7
        # elif self.id == '0000031b':
        #     self.ws = self.wb.worksheets[3]
        #     self.len = 30

    def title(self):
        if self.id == '00000146':
            self.wb2 = Workbook()
            self.ws2 = self.wb2.create_sheet(title='UDP_MSG_146', index=0)
            for i in range(2, self.len):
                self.ws2.cell(1, i).value = self.ws.cell(i, 1).value
            self.wb2.save('UDP_MSG_146.xlsx')
            self.wb2.close()
        elif self.id == '0000017f':
            self.wb2 = Workbook()
            self.ws2 = self.wb2.create_sheet(title='UDP_MSG_17F', index=0)
            for i in range(2, self.len):
                self.ws2.cell(1, i).value = self.ws.cell(i, 1).value
            self.wb2.save('UDP_MSG_17F.xlsx')
            self.wb2.close()
        # elif self.id == '00000271':
        #     self.wb2 = Workbook()
        #     self.ws2 = self.wb2.create_sheet(title='UDP_MSG_271', index=0)
        #     for i in range(2, self.len):
        #         self.ws2.cell(1, i).value = self.ws.cell(i, 1).value
        #     self.wb2.save('UDP_MSG_271.xlsx')
        #     self.wb2.close()
        # elif self.id == '0000031b':
        #     self.wb2 = Workbook()
        #     self.ws2 = self.wb2.create_sheet(title='UDP_MSG_31B', index=0)
        #     for i in range(2, self.len):
        #         self.ws2.cell(1, i).value = self.ws.cell(i, 1).value
        #     self.wb2.save('UDP_MSG_31B.xlsx')
        #     self.wb2.close()

    def cut(self):
        self.hex_lst = []
        for i in range(2, self.len):
            a = self.ws.cell(i, 3).value * 2
            b = self.ws.cell(i, 4).value * 2 + a
            self.hex_lst.append(self.msg[a:b])
        print(self.hex_lst)

    def go_int(self):
        self.int_lst = []
        for i in self.hex_lst:
            i = int(i, 16)
            self.int_lst.append(i)

    def calc(self):
        self.msg_lst = []
        for i in range(2, self.len):
            fa = self.ws.cell(i, 5).value
            of = self.ws.cell(i, 6).value
            j = self.int_lst[i - 2] * fa + of
            self.msg_lst.append(j)

    def op(self):
        if self.id == '00000146':
            self.wb2 = openpyxl.load_workbook('UDP_MSG_146.xlsx')
            self.ws2 = self.wb2.worksheets[0]
        elif self.id == '0000017f':
            self.wb2 = openpyxl.load_workbook('UDP_MSG_17F.xlsx')
            self.ws2 = self.wb2.worksheets[0]
        # elif self.id == '00000271':
        #     self.wb2 = openpyxl.load_workbook('UDP_MSG_271.xlsx')
        #     self.ws2 = self.wb2.worksheets[0]
        # elif self.id == '0000031b':
        #     self.wb2 = openpyxl.load_workbook('UDP_MSG_31B.xlsx')
        #     self.ws2 = self.wb2.worksheets[0]

    def wri(self):
        row = self.row + 2
        self.ws2.cell(row, 1).value = self.row
        for i in range(2, self.len):
            self.ws2.cell(row, i).value = self.msg_lst[i - 2]

    def save(self):
        if self.id == '00000146':
            self.wb2.save('UDP_MSG_146.xlsx')
            self.wb2.close()
        elif self.id == '0000017f':
            self.wb2.save('UDP_MSG_17F.xlsx')
            self.wb2.close()
        # elif self.id == '00000271':
        #     self.wb2.save('UDP_MSG_271.xlsx')
        #     self.wb2.close()
        # elif self.id == '0000031b':
        #     self.wb2.save('UDP_MSG_31B.xlsx')
        #     self.wb2.close()

if __name__ == '__main__':
    da_146 = '0011031900000077000001465609d51b0000000000976103d0c5009cbe03cd1a00c30503cfcd0002d31e01273c271227422777274526ef09ac5e0f76640f72051d74b300000000000000000000000000001c308c0119180098b103d0b6009ccb03cd1900c32903cfcc1e83e60f42930f83e40f2ec6018002f4f17ff8eae58000247301626902bf16050bc3'
    da_17f = '0011031e000000450000017f865c5bde000000000376f3010a0000004752955d000500aaec5c77000502c6ac0064000162730064001339bd0005051b3c1b3c1b391b3a273d273d08d20fe907dc00000000001c30f001011918'
    da_271 = '001103240000000b000002710a34753700000000001c31540101000098af98'
    da_31b = '00110339000000380000031b3dbab91d00000000001704130a0820009601a97a0000016d43016ebf0187bd00c100b000c200047e00000000008c0000500000780000000000002c1600000001'

    file = open('../function/log_146.txt', 'r', encoding='gbk', errors='ignore')
    line = file.readlines()

    row = 0
    for i in line:
        da = udp_data(i, row)

        if row == 0:
            da.cr_title()

        da.log()
        row = row + 1
