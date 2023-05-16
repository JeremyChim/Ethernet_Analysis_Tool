import openpyxl

wb = openpyxl.load_workbook('config.xlsx')
ws = wb.worksheets[3]

ls1 = ' '
ls2 = []
ls3 = []
ls4 = []
ls5 = []

for i in range(5):
    i = i + 2

    t1 = ws.cell(i, 1).value
    t2 = ws.cell(i, 2).value * 2
    t3 = ws.cell(i, 3).value * 2 + t2
    t4 = ws.cell(i, 4).value
    t5 = ws.cell(i, 5).value

    ls1 = ls1 + ',' + str(t1)
    ls2.append(t2)
    ls3.append(t3)
    ls4.append(t4)
    ls5.append(t5)

print(f"a = '{ls1}'")
print(f'b = {ls2}')
print(f'c = {ls3}')
print(f'd = {ls4}')
print(f'e = {ls5}')

