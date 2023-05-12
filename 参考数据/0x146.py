# -*- coding: utf-8 -*-
# @Time ： 2023/5/11 17:08
# @Auth ： JeremyChim
# @File ：0x146.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim

from openpyxl import Workbook
import openpyxl

path='1.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.worksheets[0]

wb2 = Workbook()
ws2 = wb2.create_sheet(title='UDP_MSG_146',index=0)

da = '00987d03cddb'
da = '0011031900000077000001465609d51b0000000000976103d0c5009cbe03cd1a00c30503cfcd0002d31e01273c271227422777274526ef09ac5e0f76640f72051d74b300000000000000000000000000001c308c0119180098b103d0b6009ccb03cd1900c32903cfcc1e83e60f42930f83e40f2ec6018002f4f17ff8eae58000247301626902bf16050bc3'
da = da[40:]

for i in range(2,43):
    ws2.cell(1, i).value = ws.cell(i, 1).value

for i in range(2,43):
    s1 = ws.cell(i, 3).value
    s1 = s1*2

    s2 = ws.cell(i, 4).value
    s2 = s1 + s2*2

    da1 = da[s1:s2]
    da2 = int(da1,16)

    a = ws.cell(i, 5).value
    b = ws.cell(i, 6).value
    re = da2 * a + b

    ws2.cell(2, 1).value = 0
    ws2.cell(2, i).value = re

wb2.save('UDP_MSG_146.csv')
wb2.close()



