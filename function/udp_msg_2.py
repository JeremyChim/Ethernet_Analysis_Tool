# -*- coding: utf-8 -*-
# @Time ： 2023/5/11 17:08
# @Auth ： JeremyChim
# @File ：udp_msg_2.py
# @IDE ：PyCharm
# @Github ：https://github.com/JeremyChim

from openpyxl import Workbook
import openpyxl

class udp_data():

    def __init__(self, udp_data:list):
        self.l = len(udp_data)
        self.id_lst = []
        self.msg_lst = []

        for i in range(self.l):
            data = udp_data[i]
            id = data[16:24]
            msg = data[40:]
            self.id_lst.append(id)
            self.msg_lst.append(msg)

        print(self.id_lst)
        print(self.msg_lst)

    def config(self):

        self.ws_lst = []
        self.len_lst = []

        for i in range(self.l):
            if self.id_lst[i] == '00000146':
                ws = 0
                len = 43
                self.ws_lst.append(ws)
                self.len_lst.append(len)

        print(self.ws_lst)
        print(self.len_lst)

    def cut(self):
        wb = openpyxl.load_workbook(filename='config.xlsx')
        lst = []
        for i in range(self.l):
            i = self.ws_lst[i]
            ws = wb.worksheets[i]
            len = self.len_lst[i]
            msg = self.msg_lst[i]
            print(i, ws, len, msg)

            for j in range(2, len):
                a = ws.cell(j, 3).value
                b = ws.cell(j, 4).value * 2 + a
                lst.append(msg[a:b])

        # for i in range(2, self.len):
        #     a = self.ws.cell(i, 3).value * 2
        #     b = self.ws.cell(i, 4).value * 2 + a
        #     self.hex_lst.append(self.msg[a:b])

    def go_int(self):
        self.int_lst = []
        for i in self.hex_lst:
            i = int(i, 16)
            self.int_lst.append(i)

    def calc(self):
        self.msg_lst = []
        for i in range(2, self.len):
            fa = self.ws.cell(i, 5).value
            of = self.ws.cell(i, 6).value
            j = self.int_lst[i - 2] * fa + of
            self.msg_lst.append(j)

    def title(self):
        if self.id == '00000146':
            self.wb2 = Workbook()
            self.ws2 = self.wb2.create_sheet(title='UDP_MSG_146', index=0)
            for i in range(2, self.len):
                self.ws2.cell(1, i).value = self.ws.cell(i, 1).value

    def save(self):
            self.ws2.cell(2, 1).value = 0
            for i in range(2, self.len):
                self.ws2.cell(2, i).value = self.msg_lst[i - 2]
            self.wb2.save('UDP_MSG_146.csv')
            self.wb2.close()

    def log(self):
        self.config()
        self.cut()
        # self.go_int()
        # self.calc()
        # self.title()
        # self.save()

if __name__ == '__main__':
    da = '0011031900000077000001465609d51b0000000000976103d0c5009cbe03cd1a00c30503cfcd0002d31e01273c271227422777274526ef09ac5e0f76640f72051d74b300000000000000000000000000001c308c0119180098b103d0b6009ccb03cd1900c32903cfcc1e83e60f42930f83e40f2ec6018002f4f17ff8eae58000247301626902bf16050bc3'
    da = [da,da]
    # print(len(da))
    da = udp_data(da)
    da.log()